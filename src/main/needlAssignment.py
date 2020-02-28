from string import ascii_uppercase
from openpyxl import load_workbook, Workbook
import os


class NeedlAssignment:

    def __init__(self, file):
        self.operators = ['-', '+', '/', '*', '%', '(', ')']
        self.methods = ['SUM', 'AVERAGE']
        self.res = dict()
        self.wb = load_workbook(filename=file)
        self.wb_data = load_workbook(filename=file, data_only=True)

    # def separate(self, val):
    #     sheet_name = ''
    #     if '!' in val:
    #         sheet_name = val.replace('=', '').split('!')[0].replace('\'', '')
    #         val = str(val.replace('=', '').split('!')[1])
    #     cell_no = val.replace('=', '')
    #     return sheet_name, cell_no

    def absoluteValue(self, sheet_name, cell_no):
        val = self.wb[sheet_name][cell_no].value
        while type(val) == str and val.startswith('='):
            if '!' in val:
                sheet_name = val.replace('=', '').split('!')[0].replace('\'', '')
                val = str(val.replace('=', '').split('!')[1])
            cell_no = val.replace('=', '')
            tmp = ''
            if '&' in cell_no:
                cell_no = cell_no.split('&')[0]
                try:
                    tmp = cell_no.split('&')[1].replace('"', '')
                except:
                    pass
            val = self.wb[sheet_name][cell_no].value + tmp
        j = 1
        while type(val) != str:
            row = self.wb[sheet_name][cell_no].row
            col = self.wb[sheet_name][cell_no].column - j
            val = self.wb[sheet_name].cell(row=row, column=col).value
            j += 1
        return val

    def formula(self, value):
        res = ''
        i = 0
        while i < len(value):
            sheet_name = 'SA-Ratios'
            cell_no = ''
            if value[i] in self.operators:
                res += value[i]
                i += 1
            # for substring startswith '\'' the value of item is in different sheet.
            # find out the sheet name for the string and then the cell no and cell value
            elif value[i] == '\'':
                i += 1
                sheet_name = ''
                # finding sheet name
                while i < len(value) and value[i] != '\'':
                    sheet_name += value[i]
                    i += 1
                i += 2
                # finding cell no
                while i < len(value) and (value[i].isalnum() or value[i] == '$'):
                    cell_no += value[i]
                    i += 1
                # need to generalize it
                if sheet_name == 'SA-Ratios':
                    row = self.wb[sheet_name][cell_no].row
                    val = self.wb_data[sheet_name].cell(row=row, column=2).value
                else:
                    val = self.absoluteValue(sheet_name, cell_no)
                res += val
            # finding methods or cell no for the same sheet
            elif value[i].isalnum():
                tmp = ''
                while i < len(value) and (value[i].isalnum() or value[i] == '$'):
                    tmp += value[i]
                    i += 1
                # finding sum or average key words in the string
                if tmp in self.methods:
                    res += tmp
                    res += value[i]
                    i += 1
                    s_name = 'SA-Ratios'
                    c_no = list()
                    # collecting the key word
                    while value[i] != ')':
                        if value[i] == '\'':
                            i += 1
                            s_name = ''
                            while value[i] != ')' and value[i] != '\'':
                                s_name += value[i]
                                i += 1
                            i += 2

                        cols = ''
                        while value[i] != ')':
                            cols += value[i]
                            i += 1
                        c_no = cols.split(':')

                    # aggregation over columns
                    if self.wb[s_name][c_no[-1]].column - self.wb[s_name][c_no[0]].column == 0:
                        row_count = abs(self.wb[s_name][c_no[-1]].row - self.wb[s_name][c_no[0]].row)
                        start_row = min(self.wb[s_name][c_no[-1]].row, self.wb[s_name][c_no[0]].row)
                        for k in range(row_count):
                            if s_name == 'SA-Ratios':
                                res += (self.wb_data[s_name].cell(row=start_row + k, column=2).value + ', ')

                            else:
                                j = 1
                                val = ''
                                while not val or type(val) != str:
                                    val = self.wb_data[s_name].cell(row=start_row + k, column=j).value
                                    j += 1
                                res += (val + ', ')
                                res = res[:-1]
                    # aggregation over rows
                    else:
                        if s_name == 'SA-Ratios':
                            row = self.wb[s_name][c_no[0]].row
                            col_count = self.wb[s_name][c_no[-1]].column - self.wb[s_name][c_no[0]].column
                            res += 'last ' + str(col_count + 1) + ' items of '
                            res += self.wb_data[s_name].cell(row=row, column=2).value

                        else:
                            j = 1
                            col_count = self.wb[s_name][c_no[-1]].column - self.wb[s_name][c_no[0]].column
                            val = ''
                            row = self.wb[s_name][c_no[0]].row
                            while not val or type(val) != str:
                                val = self.wb_data[s_name].cell(row=row, column=j).value
                                j += 1
                            res += 'last ' + str(col_count + 1) + ' items of '
                            res += val
                    res += value[i]
                    i += 1

                # if the key word is just number
                elif tmp[0].isdigit():
                    res += tmp

                # if the keyword is cell no in the same sheet
                else:
                    row = self.wb[sheet_name][tmp].row
                    val = self.wb[sheet_name].cell(row=row, column=2).value
                    while val.startswith('='):
                        tmp = ''
                        if '!' in val:
                            sheet_name = val.replace('=', '').split('!')[0].replace('\'', '')
                            val = str(val.replace('=', '').split('!')[1])
                        if '&' in val:
                            tmp = val.split('&')[1].replace('"', '')
                            val = val.split('&')[0]
                        val = self.absoluteValue(sheet_name, val.replace('=', '')) + tmp
                    res += val

            else:
                res += value[i]
                i += 1
        return res

    def execute(self):
        columns = list(ascii_uppercase)[2:19]
        ratio_sheet = self.wb['SA-Ratios']
        data_sheet = self.wb_data['SA-Ratios']
        for j, column in enumerate(columns):
            for i, cell in enumerate(ratio_sheet[column]):
                if i < 7:
                    continue
                if cell.value and type(cell.value) == str and cell.value.startswith('='):
                    key = ratio_sheet.cell(row=cell.row, column=cell.column - j - 1).value
                    if key is None:
                        continue
                    if key.startswith('='):
                        key = data_sheet.cell(row=cell.row, column=cell.column - j - 1).value
                    # while key.startswith('='):
                    #     sheet_name = 'SA-Ratios'
                    #     if '!' in key:
                    #         sheet_name = key.replace('=', '').split('!')[0].replace('\'', '')
                    #         key = str(key.replace('=', '').split('!')[1])
                    #
                    #     index = key.replace('=', '')
                    #     if '&' in index:
                    #         index = index.split('&')[0]
                    #     key = self.wb[sheet_name][index].value
                    if cell.row in self.res:
                        continue
                    self.res[cell.row] = self.formula(cell.value.replace('=', ''))

        output = '/Users/shivamchoubey/Desktop/Personal/Assignment/needlAssignment.xlsx'
        os.remove(output)
        out = Workbook()
        res_sheet = out.active
        res_sheet.title = 'Formula'
        for i, key in enumerate(sorted(self.res)):
            res_sheet.cell(column=1, row=i + 1, value=data_sheet.cell(row=int(key), column=2).value)
            res_sheet.cell(column=2, row=i + 1, value=self.res[key])
        out.save(output)


if __name__ == "__main__":
    assignment = NeedlAssignment('/Users/shivamchoubey/Desktop/Personal/Assignment/NF-SA Template 160519.xlsx')
    assignment.execute()